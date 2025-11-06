"""Excel writer module for generating DSCO-formatted inventory files."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, numbers
import logging
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Handles writing inventory data to Excel files in DSCO template format."""
    
    def __init__(self):
        """Initialize the Excel writer."""
        pass
    
    @staticmethod
    def generate_output_filename(base_name: str = "inventory", output_dir: str = ".") -> str:
        """
        Generate output filename with timestamp.
        
        Creates a filename in the format: {base_name}_YYYYMMDD_HHMMSS.xlsx
        Ensures the output directory exists.
        
        Args:
            base_name: Base name for the file (default: "inventory")
            output_dir: Directory where file will be saved (default: current directory)
            
        Returns:
            Full path to the output file
        """
        # Ensure output directory exists
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{base_name}_{timestamp}.xlsx"
        
        return str(output_path / filename)
    
    def write(self, dataframe: pd.DataFrame, output_path: str) -> None:
        """
        Write DataFrame to Excel file with DSCO template formatting.
        
        Args:
            dataframe: DataFrame with columns: SKU, Description, Warehouse, Quantity, Date
            output_path: Path where the Excel file should be saved
            
        Raises:
            Exception: If writing fails
        """
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventory"
            
            # Add column headers matching DSCO template
            headers = ['SKU', 'Description', 'Warehouse', 'Quantity', 'Date']
            ws.append(headers)
            
            # Make headers bold
            for cell in ws[1]:
                cell.font = Font(bold=True)
            
            # Write DataFrame data to worksheet
            for r_idx, row in enumerate(dataframe.itertuples(index=False), start=2):
                ws.append(list(row))
            
            # Apply formatting
            self._format_columns(ws)
            
            # Save the workbook
            wb.save(output_path)
            logger.info(f"Successfully wrote {len(dataframe)} rows to {output_path}")
            
        except Exception as e:
            error_msg = f"Failed to write Excel file: {e}"
            logger.error(error_msg)
            raise Exception(error_msg)
    
    def _format_columns(self, worksheet) -> None:
        """
        Apply formatting to worksheet columns for readability.
        
        Sets column widths and formats date and number columns.
        
        Args:
            worksheet: openpyxl worksheet object
        """
        # Set column widths for readability
        # Column A: SKU
        worksheet.column_dimensions['A'].width = 20
        # Column B: Description
        worksheet.column_dimensions['B'].width = 40
        # Column C: Warehouse
        worksheet.column_dimensions['C'].width = 15
        # Column D: Quantity
        worksheet.column_dimensions['D'].width = 12
        # Column E: Date
        worksheet.column_dimensions['E'].width = 15
        
        # Format number columns (Quantity - column D)
        for row in range(2, worksheet.max_row + 1):
            # Format quantity as integer
            qty_cell = worksheet.cell(row=row, column=4)
            qty_cell.number_format = '0'
            
            # Format date column (column E)
            date_cell = worksheet.cell(row=row, column=5)
            date_cell.number_format = 'yyyy-mm-dd'
